from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


EXPECTED_SHEETS = [
    "adf_pp",
    "descriptive",
    "correlation",
    "vif",
    "quantile_reg",
    "qr_pseudo_r2",
    "qr_quantile_diff",
    "diagnostics",
    "rolling_252",
    "rolling_504",
    "qlp_full",
    "qlp_2012_2019",
    "qlp_2020_2022",
    "qlp_2023_present",
    "qlp_key_results",
]

MAIN_VARS = ["d_vstoxx", "d_bund", "r_ttf", "r_brent", "r_eua"]
SUBSAMPLES = ["2012_2019", "2020_2022", "2023_present"]

TAU_COLORS = {
    0.10: "#1f4e79",
    0.25: "#1f4e79",
    0.50: "#d97706",
    0.75: "#2f6b2f",
    0.90: "#2f6b2f",
}

TAU_COLORS_QLP = {
    0.10: "#1f4e79",
    0.50: "#d97706",
    0.90: "#2f6b2f",
}

SAMPLE_COLORS = {
    "2012_2019": "#1f4e79",
    "2020_2022": "#d97706",
    "2023_present": "#2f6b2f",
}


def configure_plot_style() -> None:
    plt.rcParams["font.family"] = "DejaVu Serif"
    plt.rcParams["axes.titlesize"] = 12
    plt.rcParams["axes.labelsize"] = 11
    plt.rcParams["xtick.labelsize"] = 10
    plt.rcParams["ytick.labelsize"] = 10
    plt.rcParams["legend.fontsize"] = 9
    plt.rcParams["figure.facecolor"] = "white"
    plt.rcParams["axes.facecolor"] = "white"


def format_axis(ax, zero_line: bool = False) -> None:
    ax.grid(axis="y", linestyle="--", linewidth=0.6, alpha=0.25)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    if zero_line:
        ax.axhline(0.0, color="black", linestyle="--", linewidth=1.0)


def safe_quantile_label(q: float) -> str:
    return f"q{q:.2f}"


def ci_excludes_zero(ci_low: float, ci_high: float) -> bool:
    if pd.isna(ci_low) or pd.isna(ci_high):
        return False
    return (ci_low > 0) or (ci_high < 0)


def find_input_workbook(base_dir: Path) -> Path:
    candidates = [
        base_dir / "output" / "output_stats.xlsx",
        base_dir / "output_stats.xlsx",
    ]
    for c in candidates:
        if c.exists():
            return c
    raise FileNotFoundError("Could not find output_stats.xlsx in expected locations.")


def ensure_output_dirs(base_dir: Path) -> Dict[str, Path]:
    final_dir = base_dir / "final_outputs"
    tables_dir = final_dir / "tables"
    figures_dir = final_dir / "figures"
    excel_tables_dir = final_dir / "excel_tables"
    for d in [final_dir, tables_dir, figures_dir, excel_tables_dir]:
        d.mkdir(parents=True, exist_ok=True)
    return {
        "final": final_dir,
        "tables": tables_dir,
        "figures": figures_dir,
        "excel_tables": excel_tables_dir,
    }


def read_sheets(input_path: Path, expected_sheets: List[str]) -> Tuple[Dict[str, pd.DataFrame], List[str], List[str]]:
    warnings_list: List[str] = []
    sheets: Dict[str, pd.DataFrame] = {}
    used_sheets: List[str] = []

    xls = pd.ExcelFile(input_path)
    available = set(xls.sheet_names)

    for s in expected_sheets:
        if s in available:
            sheets[s] = pd.read_excel(input_path, sheet_name=s)
            used_sheets.append(s)
        else:
            warnings_list.append(f"Missing sheet: {s}")

    # Optional QR subsample sheets (naming can vary).
    for s in xls.sheet_names:
        if s.startswith("qr_") and s not in sheets:
            sheets[s] = pd.read_excel(input_path, sheet_name=s)
            used_sheets.append(s)

    return sheets, used_sheets, warnings_list


def save_table_excel(df: pd.DataFrame, output_path: Path) -> None:
    df.to_excel(output_path, index=False)


def render_table_png(df: pd.DataFrame, title: str, output_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(10, max(2.5, 0.4 * len(df) + 1.8)))
    ax.axis("off")
    tbl = ax.table(
        cellText=df.values,
        colLabels=df.columns,
        cellLoc="center",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(9)
    tbl.scale(1, 1.2)
    ax.set_title(title, fontsize=12, pad=12)
    fig.tight_layout()
    fig.savefig(output_path, dpi=300)
    plt.close(fig)


def build_stationarity_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "adf_pp" not in sheets:
        return pd.DataFrame()
    df = sheets["adf_pp"].copy()
    cols = ["series", "adf_stat", "adf_pvalue", "pp_stat", "pp_pvalue"]
    df = df[cols].copy()
    df["adf_stat"] = df["adf_stat"].round(4)
    df["pp_stat"] = df["pp_stat"].round(4)
    df["adf_pvalue"] = df["adf_pvalue"].round(5)
    df["pp_pvalue"] = df["pp_pvalue"].round(5)
    df["Stationary?"] = np.where(
        (df["adf_pvalue"] < 0.05) & (df["pp_pvalue"] < 0.05),
        "Yes",
        "No",
    )
    return df


def build_descriptive_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "descriptive" not in sheets:
        return pd.DataFrame()
    df = sheets["descriptive"].copy()
    var_col = df.columns[0]
    out = df[[var_col, "mean", "std", "min", "50%", "max", "count"]].copy()
    out = out.rename(columns={var_col: "variable", "50%": "median"})
    for c in ["mean", "std", "min", "median", "max"]:
        out[c] = out[c].round(6)
    out["count"] = out["count"].round(0).astype("Int64")
    return out


def build_correlation_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "correlation" not in sheets:
        return pd.DataFrame()
    df = sheets["correlation"].copy()
    idx_col = df.columns[0]
    out = df.rename(columns={idx_col: "variable"}).copy()
    numeric_cols = [c for c in out.columns if c != "variable"]
    out[numeric_cols] = out[numeric_cols].round(3)
    return out


def build_vif_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "vif" not in sheets:
        return pd.DataFrame()
    df = sheets["vif"].copy()
    out = df[["variable", "vif"]].copy()
    out["vif"] = out["vif"].round(3)
    out["Multicollinearity concern?"] = np.where(out["vif"] > 5, "Yes", "No")
    return out


def build_qr_full_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "quantile_reg" not in sheets:
        return pd.DataFrame()
    df = sheets["quantile_reg"].copy()
    quantiles = [0.10, 0.25, 0.50, 0.75, 0.90]
    out_rows = []
    for var in MAIN_VARS:
        row = {"variable": var}
        sub = df[df["param"] == var]
        for q in quantiles:
            sq = sub[np.isclose(sub["quantile"], q)]
            if sq.empty:
                row[f"{safe_quantile_label(q)} coef"] = ""
                row[f"{safe_quantile_label(q)} CI"] = ""
                continue
            coef = float(sq["coef"].iloc[0])
            ci_low = float(sq["ci_low"].iloc[0])
            ci_high = float(sq["ci_high"].iloc[0])
            star = "*" if ci_excludes_zero(ci_low, ci_high) else ""
            row[f"{safe_quantile_label(q)} coef"] = f"{coef:.6f}{star}"
            row[f"{safe_quantile_label(q)} CI"] = f"[{ci_low:.6f}, {ci_high:.6f}]"
        out_rows.append(row)
    return pd.DataFrame(out_rows)


def build_qr_pseudo_r2_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "qr_pseudo_r2" not in sheets:
        return pd.DataFrame()
    df = sheets["qr_pseudo_r2"].copy()
    df["quantile"] = df["quantile"].round(2)
    df["pseudo_r2"] = df["pseudo_r2"].round(4)
    return df[["quantile", "pseudo_r2"]]


def build_qr_quantile_diff_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "qr_quantile_diff" not in sheets:
        return pd.DataFrame()
    df = sheets["qr_quantile_diff"].copy()
    df = df[df["param"] != "const"].copy()
    out = df[["q1", "q2", "param", "diff_mean", "pvalue"]].copy()
    out["q1"] = out["q1"].round(2)
    out["q2"] = out["q2"].round(2)
    out["diff_mean"] = out["diff_mean"].round(6)
    out["pvalue"] = out["pvalue"].round(5)
    out["Significant difference?"] = np.where(out["pvalue"] < 0.05, "Yes", "No")
    return out


def build_diagnostics_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "diagnostics" not in sheets:
        return pd.DataFrame()
    df = sheets["diagnostics"].copy()
    out = df[["quantile", "ljungbox_pvalue", "arch_pvalue", "jb_pvalue"]].copy()
    out["ljungbox_pvalue"] = out["ljungbox_pvalue"].round(5)
    out["arch_pvalue"] = out["arch_pvalue"].round(5)
    out["jb_pvalue"] = out["jb_pvalue"].round(5)
    out["Residual autocorrelation?"] = np.where(out["ljungbox_pvalue"] < 0.05, "Yes", "No")
    out["ARCH effects?"] = np.where(out["arch_pvalue"] < 0.05, "Yes", "No")
    out["Normality rejected?"] = np.where(out["jb_pvalue"] < 0.05, "Yes", "No")
    return out


def build_subsample_summary_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    qr_sample_sheets = [s for s in sheets if s.startswith("qr_") and s not in {"qr_pseudo_r2", "qr_quantile_diff"}]
    for sname in sorted(qr_sample_sheets):
        if sname in {"quantile_reg"}:
            continue
        sample_label = sname.replace("qr_", "")
        if sample_label == "pseudo_r2" or sample_label == "quantile_diff":
            continue
        df = sheets[sname]
        required_cols = {"quantile", "param", "coef", "ci_low", "ci_high"}
        if not required_cols.issubset(set(df.columns)):
            continue
        for var in MAIN_VARS:
            sub = df[df["param"] == var].copy()
            if sub.empty:
                continue
            sub["significant"] = sub.apply(
                lambda r: ci_excludes_zero(r["ci_low"], r["ci_high"]),
                axis=1,
            )
            sig = sub[sub["significant"]].copy()
            if sig.empty:
                sig_quantiles = "None"
                sign_pattern = "n.s."
                interpretation = "No statistically significant effect across quantiles."
            else:
                sig_quantiles = ", ".join([safe_quantile_label(q) for q in sig["quantile"]])
                signs = np.sign(sig["coef"]).tolist()
                if all(v > 0 for v in signs):
                    sign_pattern = "Positive"
                    interpretation = "Positive effect in significant quantiles."
                elif all(v < 0 for v in signs):
                    sign_pattern = "Negative"
                    interpretation = "Negative effect in significant quantiles."
                else:
                    sign_pattern = "Mixed"
                    interpretation = "Sign changes across significant quantiles."
            rows.append({
                "sample_label": sample_label,
                "variable": var,
                "significant_quantiles": sig_quantiles,
                "sign_pattern": sign_pattern,
                "short_interpretation": interpretation,
            })
    return pd.DataFrame(rows)


def build_qlp_full_table(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if "qlp_key_results" not in sheets:
        return pd.DataFrame()
    df = sheets["qlp_key_results"].copy()
    df = df[df["sample_label"] == "full"].copy()
    if df.empty:
        return pd.DataFrame()
    horizons = [1, 5, 10, 20]
    quantiles = [0.10, 0.50, 0.90]
    rows = []
    for var in MAIN_VARS:
        for h in horizons:
            row = {"variable": var, "horizon": h}
            sub = df[(df["param"] == var) & (df["horizon"] == h)]
            for q in quantiles:
                sq = sub[np.isclose(sub["quantile"], q)]
                if sq.empty:
                    row[f"{safe_quantile_label(q)} coef"] = ""
                    row[f"{safe_quantile_label(q)} CI"] = ""
                    continue
                coef = float(sq["coef"].iloc[0])
                ci_low = float(sq["ci_low"].iloc[0])
                ci_high = float(sq["ci_high"].iloc[0])
                star = "*" if ci_excludes_zero(ci_low, ci_high) else ""
                row[f"{safe_quantile_label(q)} coef"] = f"{coef:.6f}{star}"
                row[f"{safe_quantile_label(q)} CI"] = f"[{ci_low:.6f}, {ci_high:.6f}]"
            rows.append(row)
    return pd.DataFrame(rows)


def build_qlp_subsample_tables(sheets: Dict[str, pd.DataFrame]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if "qlp_key_results" not in sheets:
        return pd.DataFrame(), pd.DataFrame()
    df = sheets["qlp_key_results"].copy()
    sub = df[df["sample_label"].isin(SUBSAMPLES)].copy()
    if sub.empty:
        return pd.DataFrame(), pd.DataFrame()

    detailed = sub[[
        "sample_label",
        "param",
        "horizon",
        "quantile",
        "coef",
        "ci_low",
        "ci_high",
    ]].copy()
    detailed = detailed.rename(columns={"param": "variable"})
    detailed["significant"] = detailed.apply(
        lambda r: "Yes" if ci_excludes_zero(r["ci_low"], r["ci_high"]) else "No",
        axis=1,
    )
    detailed["coef"] = detailed["coef"].round(6)
    detailed["ci_low"] = detailed["ci_low"].round(6)
    detailed["ci_high"] = detailed["ci_high"].round(6)

    rows = []
    for sample in SUBSAMPLES:
        for var in MAIN_VARS:
            ss = detailed[(detailed["sample_label"] == sample) & (detailed["variable"] == var)]
            sig = ss[ss["significant"] == "Yes"]
            if sig.empty:
                marker = "None"
                interp = "No significant dynamic responses."
            else:
                marker = "; ".join(
                    [
                        f"h{int(r['horizon'])}-{safe_quantile_label(float(r['quantile']))}"
                        for _, r in sig.sort_values(["horizon", "quantile"]).iterrows()
                    ]
                )
                interp = "Significant responses present at listed horizon-quantile pairs."
            rows.append({
                "sample_label": sample,
                "variable": var,
                "significant_horizon_quantiles": marker,
                "short_interpretation": interp,
            })
    summary = pd.DataFrame(rows)
    return detailed, summary


def plot_index_levels(base_dir: Path, figures_dir: Path, warnings_list: List[str], generated_figures: List[str]) -> None:
    input_candidates = [
        base_dir / "date_simp_concat.xlsx",
        base_dir / "date simp.xlsx",
    ]
    source_path = None
    for c in input_candidates:
        if c.exists():
            source_path = c
            break
    if source_path is None:
        warnings_list.append("Figure index levels skipped: input levels file not found.")
        return

    try:
        df = pd.read_excel(source_path)
    except Exception as exc:
        warnings_list.append(f"Figure index levels skipped: could not read file ({exc}).")
        return

    required_cols = {"Date", "STOXX 600 ESG-X", "STOXX 600"}
    if not required_cols.issubset(set(df.columns)):
        warnings_list.append("Figure index levels skipped: required columns not found.")
        return

    temp = df[["Date", "STOXX 600 ESG-X", "STOXX 600"]].copy()
    temp["Date"] = pd.to_datetime(temp["Date"], errors="coerce", dayfirst=True)
    temp = temp.dropna(subset=["Date"])
    if temp.empty:
        warnings_list.append("Figure index levels skipped: no valid dates.")
        return

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.plot(temp["Date"], temp["STOXX 600 ESG-X"], color="#1f4e79", linewidth=1.8, label="STOXX 600 ESG-X")
    ax.plot(temp["Date"], temp["STOXX 600"], color="#4b5563", linewidth=1.8, label="STOXX 600")
    format_axis(ax, zero_line=False)
    ax.set_title("Index Levels: STOXX 600 ESG-X vs STOXX 600")
    ax.set_xlabel("Date")
    ax.set_ylabel("Index level")
    ax.legend(frameon=False, loc="best")
    fig.tight_layout()
    out_path = figures_dir / "figure_index_levels.png"
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    generated_figures.append(out_path.name)


def plot_qr_figures(sheets: Dict[str, pd.DataFrame], figures_dir: Path, warnings_list: List[str], generated_figures: List[str]) -> None:
    if "quantile_reg" not in sheets:
        warnings_list.append("QR figures skipped: sheet quantile_reg missing.")
        return
    df = sheets["quantile_reg"].copy()
    quantiles = [0.10, 0.25, 0.50, 0.75, 0.90]

    def draw_qr(ax, var: str) -> None:
        sub = df[df["param"] == var].copy().sort_values("quantile")
        if sub.empty:
            ax.set_visible(False)
            return
        x = sub["quantile"].to_numpy()
        y = sub["coef"].to_numpy()
        low = sub["ci_low"].to_numpy()
        high = sub["ci_high"].to_numpy()
        ax.plot(x, y, color="#374151", linewidth=1.8)
        ax.fill_between(x, low, high, color="#9ca3af", alpha=0.18)
        for q, yy in zip(x, y):
            ax.scatter([q], [yy], color=TAU_COLORS.get(round(float(q), 2), "#374151"), s=36, zorder=3)
        format_axis(ax, zero_line=True)
        ax.set_title(var)
        ax.set_xlabel("Quantile")
        ax.set_ylabel("Coefficient")
        ax.set_xticks(quantiles)

    # Individual QR figures
    for var in MAIN_VARS:
        fig, ax = plt.subplots(figsize=(8, 5))
        draw_qr(ax, var)
        fig.suptitle("Quantile Regression Coefficients across the Conditional Distribution of the ESG Spread")
        fig.tight_layout()
        out_path = figures_dir / f"figure_qr_{var}.png"
        fig.savefig(out_path, dpi=300)
        plt.close(fig)
        generated_figures.append(out_path.name)

    # Combined QR panel
    fig, axes = plt.subplots(2, 3, figsize=(12, 8))
    axes_flat = axes.flatten()
    for i, var in enumerate(MAIN_VARS):
        draw_qr(axes_flat[i], var)
    axes_flat[-1].axis("off")
    fig.suptitle("Quantile Regression Coefficients across the Conditional Distribution of the ESG Spread")
    fig.tight_layout()
    out_path = figures_dir / "figure_qr_combined.png"
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    generated_figures.append(out_path.name)


def plot_rolling_figures(sheets: Dict[str, pd.DataFrame], figures_dir: Path, warnings_list: List[str], generated_figures: List[str]) -> None:
    if "rolling_504" not in sheets:
        warnings_list.append("Rolling figures skipped: sheet rolling_504 missing.")
        return

    df = sheets["rolling_504"].copy()
    if "end_date" in df.columns:
        df["end_date"] = pd.to_datetime(df["end_date"], errors="coerce")
    else:
        warnings_list.append("Rolling figures skipped: end_date column missing.")
        return

    vars_to_plot = ["r_brent", "d_vstoxx"]
    taus = [0.10, 0.50, 0.90]

    for var in vars_to_plot:
        fig, ax = plt.subplots(figsize=(8, 5))
        sub = df[df["param"] == var].copy()
        if sub.empty:
            warnings_list.append(f"Rolling figure skipped for {var}: no data.")
            plt.close(fig)
            continue
        for q in taus:
            sq = sub[np.isclose(sub["quantile"], q)].sort_values("end_date")
            ax.plot(
                sq["end_date"],
                sq["coef"],
                color=TAU_COLORS_QLP[q],
                linewidth=1.5,
                label=f"tau={q:.2f}",
            )
        format_axis(ax, zero_line=True)
        ax.set_title(f"Rolling Quantile Coefficients (504-day): {var}")
        ax.set_xlabel("Window end date")
        ax.set_ylabel("Coefficient")
        ax.legend(frameon=False, loc="best")
        fig.tight_layout()
        out_path = figures_dir / f"figure_rolling_504_{var}.png"
        fig.savefig(out_path, dpi=300)
        plt.close(fig)
        generated_figures.append(out_path.name)

    # Combined rolling panel
    fig, axes = plt.subplots(2, 1, figsize=(10, 7), sharex=True)
    for ax, var in zip(axes, vars_to_plot):
        sub = df[df["param"] == var].copy()
        for q in taus:
            sq = sub[np.isclose(sub["quantile"], q)].sort_values("end_date")
            ax.plot(
                sq["end_date"],
                sq["coef"],
                color=TAU_COLORS_QLP[q],
                linewidth=1.5,
                label=f"tau={q:.2f}",
            )
        format_axis(ax, zero_line=True)
        ax.set_title(var)
        ax.set_ylabel("Coefficient")
    axes[-1].set_xlabel("Window end date")
    axes[0].legend(frameon=False, loc="best")
    fig.suptitle("Rolling Quantile Coefficients for Brent and VSTOXX")
    fig.tight_layout()
    out_path = figures_dir / "figure_rolling_combined.png"
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    generated_figures.append(out_path.name)


def plot_qlp_full_figures(sheets: Dict[str, pd.DataFrame], figures_dir: Path, warnings_list: List[str], generated_figures: List[str]) -> None:
    if "qlp_key_results" not in sheets:
        warnings_list.append("QLP figures skipped: sheet qlp_key_results missing.")
        return
    df = sheets["qlp_key_results"].copy()
    df = df[df["sample_label"] == "full"].copy()
    if df.empty:
        warnings_list.append("QLP full figures skipped: no full-sample rows.")
        return

    horizons = [1, 5, 10, 20]
    taus = [0.10, 0.50, 0.90]

    def draw_qlp(ax, var: str) -> None:
        sub = df[df["param"] == var].copy()
        if sub.empty:
            ax.set_visible(False)
            return
        for q in taus:
            sq = sub[np.isclose(sub["quantile"], q)].sort_values("horizon")
            ax.plot(
                sq["horizon"],
                sq["coef"],
                marker="o",
                linewidth=1.8,
                color=TAU_COLORS_QLP[q],
                label=f"tau={q:.2f}",
            )
            ax.fill_between(
                sq["horizon"].to_numpy(),
                sq["ci_low"].to_numpy(),
                sq["ci_high"].to_numpy(),
                color=TAU_COLORS_QLP[q],
                alpha=0.16,
            )
        format_axis(ax, zero_line=True)
        ax.set_title(var)
        ax.set_xlabel("Horizon")
        ax.set_ylabel("QLP coefficient")
        ax.set_xticks(horizons)

    for var in MAIN_VARS:
        fig, ax = plt.subplots(figsize=(8, 5))
        draw_qlp(ax, var)
        ax.legend(frameon=False, loc="best")
        fig.suptitle("Dynamic Quantile Local Projection Responses of the ESG Spread")
        fig.tight_layout()
        out_path = figures_dir / f"figure_qlp_full_{var}.png"
        fig.savefig(out_path, dpi=300)
        plt.close(fig)
        generated_figures.append(out_path.name)

    fig, axes = plt.subplots(2, 3, figsize=(12, 8))
    axes_flat = axes.flatten()
    for i, var in enumerate(MAIN_VARS):
        draw_qlp(axes_flat[i], var)
    axes_flat[-1].axis("off")
    handles, labels = axes_flat[0].get_legend_handles_labels()
    if handles:
        axes_flat[0].legend(handles, labels, frameon=False, loc="best")
    fig.suptitle("Dynamic Quantile Local Projection Responses of the ESG Spread")
    fig.tight_layout()
    out_path = figures_dir / "figure_qlp_full_combined.png"
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    generated_figures.append(out_path.name)


def plot_qlp_subsample_figures(sheets: Dict[str, pd.DataFrame], figures_dir: Path, warnings_list: List[str], generated_figures: List[str]) -> None:
    if "qlp_key_results" not in sheets:
        warnings_list.append("QLP subsample figures skipped: sheet qlp_key_results missing.")
        return
    df = sheets["qlp_key_results"].copy()
    df = df[df["sample_label"].isin(SUBSAMPLES)].copy()
    if df.empty:
        warnings_list.append("QLP subsample figures skipped: no subsample rows.")
        return

    vars_to_plot = ["r_brent", "d_vstoxx"]
    taus = [0.10, 0.50, 0.90]
    horizons = [1, 5, 10, 20]

    for var in vars_to_plot:
        fig, axes = plt.subplots(1, 3, figsize=(14, 4), sharey=True)
        for ax, sample in zip(axes, SUBSAMPLES):
            sub = df[(df["sample_label"] == sample) & (df["param"] == var)].copy()
            for q in taus:
                sq = sub[np.isclose(sub["quantile"], q)].sort_values("horizon")
                ax.plot(
                    sq["horizon"],
                    sq["coef"],
                    marker="o",
                    linewidth=1.5,
                    color=TAU_COLORS_QLP[q],
                    label=f"tau={q:.2f}",
                )
                ax.fill_between(
                    sq["horizon"].to_numpy(),
                    sq["ci_low"].to_numpy(),
                    sq["ci_high"].to_numpy(),
                    color=TAU_COLORS_QLP[q],
                    alpha=0.14,
                )
            format_axis(ax, zero_line=True)
            ax.set_title(sample)
            ax.set_xlabel("Horizon")
            ax.set_xticks(horizons)
        axes[0].set_ylabel("QLP coefficient")
        handles, labels = axes[0].get_legend_handles_labels()
        if handles:
            axes[-1].legend(handles, labels, frameon=False, loc="best")
        fig.suptitle(f"Subsample Evidence: QLP Responses for {var}")
        fig.tight_layout()
        out_path = figures_dir / f"figure_qlp_subsamples_{var}.png"
        fig.savefig(out_path, dpi=300)
        plt.close(fig)
        generated_figures.append(out_path.name)


def apply_table_workbook_format(workbook_path: Path, notes: Dict[str, str]) -> None:
    wb = load_workbook(workbook_path)

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    diag_fill = PatternFill(fill_type="solid", fgColor="EEF3FA")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="BFC5D2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_col = ws.max_column
        max_row = ws.max_row

        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = center
                cell.border = border
                if isinstance(cell.value, (int, np.integer)):
                    cell.number_format = "0"
                elif isinstance(cell.value, (float, np.floating)):
                    cell.number_format = "0.000000"

        if ws.title == "correlation":
            # Highlight correlation diagonal cells (excluding the variable column).
            for r in range(2, ws.max_row + 1):
                var_name = ws.cell(row=r, column=1).value
                if var_name is None:
                    continue
                for c in range(2, ws.max_column + 1):
                    col_name = ws.cell(row=1, column=c).value
                    if col_name == var_name:
                        ws.cell(row=r, column=c).fill = diag_fill

        for c in range(1, max_col + 1):
            letter = get_column_letter(c)
            max_len = 0
            for r in range(1, ws.max_row + 1):
                val = ws.cell(row=r, column=c).value
                val_str = "" if val is None else str(val)
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 50)

        note = notes.get(ws.title)
        if note:
            note_row = ws.max_row + 2
            ws.cell(row=note_row, column=1, value=f"Note: {note}")
            ws.cell(row=note_row, column=1).font = Font(italic=True, size=10, color="4B5563")

    wb.save(workbook_path)


def main() -> None:
    configure_plot_style()

    base_dir = Path(__file__).resolve().parent
    output_dirs = ensure_output_dirs(base_dir)
    input_path = find_input_workbook(base_dir)

    warnings_list: List[str] = []
    generated_tables: List[str] = []
    generated_figures: List[str] = []
    used_sheets: List[str] = []

    sheets, used_sheets, read_warnings = read_sheets(input_path, EXPECTED_SHEETS)
    warnings_list.extend(read_warnings)

    table_builders = [
        ("tables_stationarity.xlsx", "stationarity", build_stationarity_table(sheets)),
        ("table_descriptive_stats.xlsx", "descriptive_stats", build_descriptive_table(sheets)),
        ("table_correlation.xlsx", "correlation", build_correlation_table(sheets)),
        ("table_vif.xlsx", "vif", build_vif_table(sheets)),
        ("table_qr_full.xlsx", "qr_full", build_qr_full_table(sheets)),
        ("table_qr_pseudor2.xlsx", "qr_pseudor2", build_qr_pseudo_r2_table(sheets)),
        ("table_qr_quantile_diff.xlsx", "qr_quantile_diff", build_qr_quantile_diff_table(sheets)),
        ("table_diagnostics.xlsx", "diagnostics", build_diagnostics_table(sheets)),
        ("table_subsample_summary.xlsx", "subsample_summary", build_subsample_summary_table(sheets)),
        ("table_qlp_full.xlsx", "qlp_full", build_qlp_full_table(sheets)),
    ]

    qlp_sub_detailed, qlp_sub_summary = build_qlp_subsample_tables(sheets)
    table_builders.extend([
        ("table_qlp_subsamples_detailed.xlsx", "qlp_subsamples_detailed", qlp_sub_detailed),
        ("table_qlp_subsamples_summary.xlsx", "qlp_subsamples_summary", qlp_sub_summary),
    ])

    formatted_tables: Dict[str, pd.DataFrame] = {}

    for filename, sheet_label, df in table_builders:
        if df.empty:
            warnings_list.append(f"Table skipped ({filename}): source data unavailable or empty.")
            continue
        excel_path = output_dirs["excel_tables"] / filename
        save_table_excel(df, excel_path)
        generated_tables.append(filename)
        formatted_tables[sheet_label] = df

    # Optional stationarity PNG table.
    if "stationarity" in formatted_tables:
        png_path = output_dirs["tables"] / "tables_stationarity.png"
        render_table_png(
            formatted_tables["stationarity"],
            "Stationarity Tests",
            png_path,
        )
        generated_figures.append(png_path.name)

    # Final formatted workbook with all tables on separate sheets.
    final_tables_workbook = output_dirs["final"] / "final_tables_formatted.xlsx"
    with pd.ExcelWriter(final_tables_workbook, engine="openpyxl") as writer:
        for sheet_label, df in formatted_tables.items():
            safe_name = sheet_label[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    notes = {
        "stationarity": "Stationary? = Yes if both ADF and PP p-values are below 0.05.",
        "descriptive_stats": "Descriptive statistics for transformed econometric series.",
        "correlation": "Pairwise Pearson correlations.",
        "vif": "Multicollinearity concern marked Yes if VIF > 5.",
        "qr_full": "CI-based significance marked with * when CI excludes zero.",
        "qr_quantile_diff": "Significant difference marked Yes if p-value < 0.05.",
        "diagnostics": "Residual diagnostic flags use a 0.05 significance threshold.",
        "qlp_full": "QLP coefficients represent cumulative future spread responses by horizon.",
    }
    apply_table_workbook_format(final_tables_workbook, notes)
    generated_tables.append(final_tables_workbook.name)

    # Figures
    plot_index_levels(base_dir, output_dirs["figures"], warnings_list, generated_figures)
    plot_qr_figures(sheets, output_dirs["figures"], warnings_list, generated_figures)
    plot_rolling_figures(sheets, output_dirs["figures"], warnings_list, generated_figures)
    plot_qlp_full_figures(sheets, output_dirs["figures"], warnings_list, generated_figures)
    plot_qlp_subsample_figures(sheets, output_dirs["figures"], warnings_list, generated_figures)

    # Reporting summary text
    summary_path = output_dirs["final"] / "reporting_outputs_summary.txt"
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("Reporting Outputs Summary\n")
        f.write("=========================\n\n")
        f.write(f"Input workbook: {input_path}\n")
        f.write(f"Output root: {output_dirs['final']}\n\n")
        f.write("Used input sheets:\n")
        for s in sorted(set(used_sheets)):
            f.write(f"- {s}\n")
        f.write("\nGenerated tables:\n")
        for t in generated_tables:
            f.write(f"- {t}\n")
        f.write("\nGenerated figures:\n")
        for g in generated_figures:
            f.write(f"- {g}\n")
        f.write("\nWarnings / skipped items:\n")
        if warnings_list:
            for w in warnings_list:
                f.write(f"- {w}\n")
        else:
            f.write("- None\n")

    print("Final reporting pipeline completed.")
    print(f"Tables generated: {len(generated_tables)}")
    print(f"Figures generated: {len(generated_figures)}")
    print(f"Output folder: {output_dirs['final']}")
    if warnings_list:
        print(f"Warnings: {len(warnings_list)} (see reporting_outputs_summary.txt)")


if __name__ == "__main__":
    main()
