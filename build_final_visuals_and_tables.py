from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import patches
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_DIR = Path(__file__).resolve().parent
DATA_FILE = PROJECT_DIR / "date_simp_concat.xlsx"
RESULTS_FILE = PROJECT_DIR / "output" / "output_stats.xlsx"
OUTPUT_DIR = PROJECT_DIR / "final_visual_package"

SUBDIRS = {
    "figures_main": OUTPUT_DIR / "figures_main",
    "figures_variants": OUTPUT_DIR / "figures_variants",
    "figures_appendix": OUTPUT_DIR / "figures_appendix",
    "figures_slides": OUTPUT_DIR / "figures_slides",
    "tables_excel": OUTPUT_DIR / "tables_excel",
    "tables_png": OUTPUT_DIR / "tables_png",
    "tables_powerpoint_ready": OUTPUT_DIR / "tables_powerpoint_ready",
    "methodology_flow": OUTPUT_DIR / "methodology_flow",
}

EXPECTED_SHEETS = [
    "adf_pp",
    "descriptive",
    "correlation",
    "vif",
    "quantile_reg",
    "qr_pseudo_r2",
    "qr_quantile_diff",
    "diagnostics",
    "rolling_252",
    "rolling_504",
    "qlp_full",
    "qlp_2012_2019",
    "qlp_2020_2022",
    "qlp_2023_present",
    "qlp_key_results",
]

MAIN_VARS = ["d_vstoxx", "d_bund", "r_ttf", "r_brent", "r_eua"]
ALL_SERIES = ["spread"] + MAIN_VARS
SUBSAMPLES = ["2012_2019", "2020_2022", "2023_present"]
HORIZONS = [1, 5, 10, 20]

DISPLAY_NAMES = {
    "spread": "ESG Spread",
    "d_vstoxx": "\u0394VSTOXX",
    "d_bund": "\u0394Bund 10Y",
    "r_ttf": "TTF Return",
    "r_brent": "Brent Return",
    "r_eua": "EUA Return",
    "STOXX 600 ESG-X": "STOXX 600 ESG-X",
    "STOXX 600": "STOXX 600",
    "VSTOXX": "VSTOXX",
    "Bunds 10Y": "Bund 10Y",
    "TTF": "TTF",
    "Brent": "Brent",
    "EUA": "EUA",
}

SLUGS = {
    "d_vstoxx": "d_vstoxx",
    "d_bund": "d_bund",
    "r_ttf": "r_ttf",
    "r_brent": "r_brent",
    "r_eua": "r_eua",
}

ROLLING_SLUGS = {
    "d_vstoxx": "vstoxx",
    "d_bund": "bund",
    "r_ttf": "ttf",
    "r_brent": "brent",
    "r_eua": "eua",
}

COLORS = {
    "navy": "#26418F",
    "green": "#86BC25",
    "burgundy": "#B5121B",
    "gray": "#5B5B5B",
    "axis": "#44546A",
    "light_gray": "#D9D9D9",
    "very_light_gray": "#F2F2F2",
    "orange": "#ED7D31",
    "light_blue": "#5B9BD5",
}

Q3_COLORS = {
    0.10: COLORS["navy"],
    0.50: COLORS["green"],
    0.90: COLORS["burgundy"],
}

Q5_COLORS = {
    0.10: COLORS["navy"],
    0.25: COLORS["light_blue"],
    0.50: COLORS["green"],
    0.75: COLORS["orange"],
    0.90: COLORS["burgundy"],
}


def configure_style() -> None:
    plt.rcParams.update(
        {
            "font.family": ["Arial", "DejaVu Sans"],
            "figure.facecolor": "white",
            "axes.facecolor": "white",
            "axes.edgecolor": COLORS["axis"],
            "axes.labelcolor": COLORS["axis"],
            "axes.titlecolor": COLORS["navy"],
            "text.color": COLORS["axis"],
            "xtick.color": COLORS["axis"],
            "ytick.color": COLORS["axis"],
            "axes.titlesize": 13,
            "axes.labelsize": 11,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "legend.fontsize": 9,
            "savefig.facecolor": "white",
            "savefig.bbox": "tight",
        }
    )


def ensure_dirs() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    for d in SUBDIRS.values():
        d.mkdir(parents=True, exist_ok=True)


def add_manifest(
    manifest: List[dict],
    path: Path,
    output_type: str,
    number: str,
    description: str,
    intended_use: str,
) -> None:
    manifest.append(
        {
            "output_file": str(path.relative_to(OUTPUT_DIR)),
            "output_type": output_type,
            "figure_or_table_number": number,
            "description": description,
            "intended_use": intended_use,
        }
    )


def export_figure(
    fig,
    out_dir: Path,
    basename: str,
    manifest: List[dict],
    number: str,
    description: str,
    intended_use: str,
) -> None:
    for ext in ["png", "pdf", "svg"]:
        path = out_dir / f"{basename}.{ext}"
        if ext == "png":
            fig.savefig(path, dpi=300)
        else:
            fig.savefig(path)
        add_manifest(manifest, path, ext.upper(), number, description, intended_use)


def format_axis(ax, zero_line: bool = False) -> None:
    ax.grid(axis="y", color=COLORS["light_gray"], alpha=0.45, linewidth=0.7)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    for side in ["left", "bottom"]:
        ax.spines[side].set_color(COLORS["gray"])
        ax.spines[side].set_linewidth(0.8)
    if zero_line:
        ax.axhline(0, color="#222222", linewidth=1.0, linestyle="--", zorder=1)


def q_label(q: float) -> str:
    return f"q{q:.2f}"


def sample_label(sample: str) -> str:
    return {
        "2012_2019": "2012-2019",
        "2020_2022": "2020-2022",
        "2023_present": "2023-2026",
        "full": "Full sample",
    }.get(sample, sample)


def significant(row: pd.Series) -> bool:
    return pd.notna(row.get("ci_low")) and pd.notna(row.get("ci_high")) and (
        (row["ci_low"] > 0) or (row["ci_high"] < 0)
    )


def read_sheets(warnings: List[str]) -> Tuple[Dict[str, pd.DataFrame], List[str]]:
    sheets: Dict[str, pd.DataFrame] = {}
    used: List[str] = []
    if not RESULTS_FILE.exists():
        warnings.append(f"Missing results workbook: {RESULTS_FILE}")
        return sheets, used

    xl = pd.ExcelFile(RESULTS_FILE)
    available = set(xl.sheet_names)
    for sheet in EXPECTED_SHEETS:
        if sheet not in available:
            warnings.append(f"Missing sheet: {sheet}")
            continue
        sheets[sheet] = pd.read_excel(RESULTS_FILE, sheet_name=sheet)
        used.append(sheet)

    for sheet in xl.sheet_names:
        if (sheet.startswith("qr_") or sheet.startswith("qlp_")) and sheet not in sheets:
            sheets[sheet] = pd.read_excel(RESULTS_FILE, sheet_name=sheet)
            used.append(sheet)

    return sheets, sorted(set(used))


def read_data(warnings: List[str]) -> pd.DataFrame:
    if not DATA_FILE.exists():
        warnings.append(f"Missing data workbook: {DATA_FILE}")
        return pd.DataFrame()
    df = pd.read_excel(DATA_FILE)
    df.columns = [str(c).strip() for c in df.columns]
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True)
        df = df.dropna(subset=["Date"]).sort_values("Date")
    return df


def round_for_table(value, digits: int = 6) -> str:
    if pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def ci_text(low, high, digits: int = 6) -> str:
    if pd.isna(low) or pd.isna(high):
        return ""
    return f"[{float(low):.{digits}f}, {float(high):.{digits}f}]"


def coef_text(row: pd.Series, digits: int = 6) -> str:
    star = "*" if significant(row) else ""
    return f"{float(row['coef']):.{digits}f}{star}"


def render_table_png(
    df: pd.DataFrame,
    out_path: Path,
    title: str,
    manifest: List[dict],
    number: str,
    intended_use: str = "all",
) -> None:
    nrows, ncols = df.shape
    width = min(max(9.5, ncols * 1.25), 18)
    height = min(max(2.8, nrows * 0.36 + 1.3), 12)
    fig, ax = plt.subplots(figsize=(width, height))
    ax.axis("off")
    tbl = ax.table(
        cellText=df.astype(str).values,
        colLabels=df.columns,
        cellLoc="center",
        loc="center",
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(8.2 if ncols > 8 else 9.4)
    tbl.scale(1.0, 1.18)
    for (r, c), cell in tbl.get_celld().items():
        cell.set_edgecolor(COLORS["light_gray"])
        cell.set_linewidth(0.5)
        if r == 0:
            cell.set_facecolor(COLORS["navy"])
            cell.get_text().set_color("white")
            cell.get_text().set_weight("bold")
        else:
            cell.set_facecolor("white")
            if c == 0:
                cell.get_text().set_ha("left")
    ax.set_title(title, color=COLORS["navy"], fontsize=13, pad=12)
    fig.tight_layout()
    fig.savefig(out_path, dpi=300)
    plt.close(fig)
    add_manifest(manifest, out_path, "PNG", number, title, intended_use)


def build_table_1(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("descriptive", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    var_col = df.columns[0]
    out = df[[var_col, "mean", "std", "min", "50%", "max", "count"]].copy()
    out.columns = ["Variable", "Mean", "Std. Dev.", "Min", "Median", "Max", "Observations"]
    out["Variable"] = out["Variable"].map(DISPLAY_NAMES).fillna(out["Variable"])
    for col in ["Mean", "Std. Dev.", "Min", "Median", "Max"]:
        out[col] = out[col].map(lambda x: round_for_table(x, 6))
    out["Observations"] = out["Observations"].astype(int)
    return out


def build_table_2(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("adf_pp", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    out = df[["series", "adf_stat", "adf_pvalue", "pp_stat", "pp_pvalue"]].copy()
    out.columns = ["Variable", "ADF statistic", "ADF p-value", "PP statistic", "PP p-value"]
    out["Variable"] = out["Variable"].map(DISPLAY_NAMES).fillna(out["Variable"])
    out["Conclusion"] = np.where(
        (out["ADF p-value"] < 0.05) & (out["PP p-value"] < 0.05),
        "Stationary",
        "Check",
    )
    for col in ["ADF statistic", "PP statistic"]:
        out[col] = out[col].map(lambda x: round_for_table(x, 4))
    for col in ["ADF p-value", "PP p-value"]:
        out[col] = out[col].map(lambda x: round_for_table(x, 5))
    return out


def build_table_3(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("correlation", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    first = df.columns[0]
    out = df.rename(columns={first: "Variable"}).copy()
    out["Variable"] = out["Variable"].map(DISPLAY_NAMES).fillna(out["Variable"])
    out = out.rename(columns={c: DISPLAY_NAMES.get(c, c) for c in out.columns})
    for col in out.columns[1:]:
        out[col] = out[col].map(lambda x: round_for_table(x, 3))
    return out


def build_table_4(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("vif", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    out = df[["variable", "vif"]].copy()
    out.columns = ["Variable", "VIF"]
    out["Variable"] = out["Variable"].map(DISPLAY_NAMES).fillna(out["Variable"])
    out["Multicollinearity concern"] = np.where(out["VIF"] > 5, "Yes", "No")
    out["VIF"] = out["VIF"].map(lambda x: round_for_table(x, 3))
    return out


def build_wide_coef_table(df: pd.DataFrame, quantiles: Iterable[float], include_horizon: bool = False) -> pd.DataFrame:
    rows = []
    horizons = sorted(df["horizon"].unique()) if include_horizon else [None]
    for var in MAIN_VARS:
        for h in horizons:
            row = {"Variable": DISPLAY_NAMES[var]}
            if include_horizon:
                row["Horizon"] = int(h)
                sub = df[(df["param"] == var) & (df["horizon"] == h)]
            else:
                sub = df[df["param"] == var]
            for q in quantiles:
                sq = sub[np.isclose(sub["quantile"], q)]
                if sq.empty:
                    row[f"{q_label(q)} coef"] = ""
                    row[f"{q_label(q)} CI"] = ""
                else:
                    r = sq.iloc[0]
                    row[f"{q_label(q)} coef"] = coef_text(r)
                    row[f"{q_label(q)} CI"] = ci_text(r["ci_low"], r["ci_high"])
            rows.append(row)
    return pd.DataFrame(rows)


def build_table_5(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("quantile_reg", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    df = df[df["param"].isin(MAIN_VARS)]
    return build_wide_coef_table(df, [0.10, 0.25, 0.50, 0.75, 0.90])


def build_table_6(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("quantile_reg", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    interpretations = {
        "r_brent": "Energy-exclusion wedge.",
        "d_vstoxx": "Market-stress channel.",
        "r_ttf": "European gas shock channel.",
        "d_bund": "Weak full-sample interest-rate channel.",
        "r_eua": "Weak or insignificant carbon-price channel.",
    }
    rows = []
    for var in ["r_brent", "d_vstoxx", "r_ttf", "d_bund", "r_eua"]:
        sub = df[df["param"] == var].copy()
        sub["sig"] = sub.apply(significant, axis=1) if not sub.empty else []
        sig = sub[sub["sig"]]
        if sig.empty:
            main_sign = "Not robust"
            sig_q = "None"
        else:
            mean_coef = sig["coef"].mean()
            main_sign = "Negative" if mean_coef < 0 else "Positive"
            sig_q = ", ".join(q_label(q) for q in sig["quantile"])
        rows.append(
            {
                "Driver": DISPLAY_NAMES[var],
                "Main sign": main_sign,
                "Significant quantiles": sig_q,
                "Economic interpretation": interpretations[var],
            }
        )
    return pd.DataFrame(rows)


def build_table_7(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("qr_quantile_diff", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    df = df[df["param"].isin(MAIN_VARS)].copy()
    out = pd.DataFrame(
        {
            "Quantile pair": df["q1"].map(q_label) + " vs " + df["q2"].map(q_label),
            "Variable": df["param"].map(DISPLAY_NAMES),
            "Difference": df["diff_mean"].map(lambda x: round_for_table(x, 6)),
            "p-value": df["pvalue"].map(lambda x: round_for_table(x, 5)),
            "Significant difference?": np.where(df["pvalue"] < 0.05, "Yes", "No"),
        }
    )
    return out


def build_table_8(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("diagnostics", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    rows = []
    for _, r in df.iterrows():
        parts = []
        parts.append("No strong residual autocorrelation" if r["ljungbox_pvalue"] > 0.05 else "Residual autocorrelation")
        parts.append("ARCH effects present" if r["arch_pvalue"] < 0.05 else "No strong ARCH effects")
        parts.append("Non-normality rejected" if r["jb_pvalue"] < 0.05 else "Normality not rejected")
        rows.append(
            {
                "Quantile": q_label(r["quantile"]),
                "Ljung-Box p-value": round_for_table(r["ljungbox_pvalue"], 5),
                "ARCH-LM p-value": round_for_table(r["arch_pvalue"], 5),
                "Jarque-Bera p-value": round_for_table(r["jb_pvalue"], 5),
                "Interpretation": "; ".join(parts),
            }
        )
    return pd.DataFrame(rows)


def build_table_9(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("qlp_key_results", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    df = df[(df["sample_label"] == "full") & df["param"].isin(MAIN_VARS)]
    return build_wide_coef_table(df, [0.10, 0.50, 0.90], include_horizon=True)


def build_table_10(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    df = sheets.get("qlp_key_results", pd.DataFrame()).copy()
    if df.empty:
        return pd.DataFrame()
    rows = []
    sub = df[df["sample_label"].isin(SUBSAMPLES) & df["param"].isin(MAIN_VARS)].copy()
    for sample in SUBSAMPLES:
        for var in MAIN_VARS:
            ss = sub[(sub["sample_label"] == sample) & (sub["param"] == var)].copy()
            if ss.empty:
                continue
            ss["sig"] = ss.apply(significant, axis=1)
            sig = ss[ss["sig"]]
            if sig.empty:
                marker = "None"
                sign = "Not robust"
                interp = "No significant dynamic response."
            else:
                marker = "; ".join(
                    f"h{int(r.horizon)}-{q_label(float(r.quantile))}" for r in sig.itertuples()
                )
                sign = "Negative" if sig["coef"].mean() < 0 else "Positive"
                interp = "Significant dynamic response at listed horizon-quantile pairs."
            rows.append(
                {
                    "Subsample": sample_label(sample),
                    "Variable": DISPLAY_NAMES[var],
                    "Significant horizons/quantiles": marker,
                    "General sign pattern": sign,
                    "Interpretation": interp,
                }
            )
    return pd.DataFrame(rows)


def build_tables(sheets: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    return {
        "T1_Descriptive": build_table_1(sheets),
        "T2_Stationarity": build_table_2(sheets),
        "T3_Correlation": build_table_3(sheets),
        "T4_VIF": build_table_4(sheets),
        "T5_QR_Full": build_table_5(sheets),
        "T6_Main_Findings": build_table_6(sheets),
        "T7_Quantile_Diff": build_table_7(sheets),
        "T8_Diagnostics": build_table_8(sheets),
        "T9_QLP_Full": build_table_9(sheets),
        "T10_QLP_Sub_Summary": build_table_10(sheets),
    }


def export_tables(
    tables: Dict[str, pd.DataFrame],
    manifest: List[dict],
    warnings: List[str],
) -> List[str]:
    generated: List[str] = []
    workbook_path = SUBDIRS["tables_excel"] / "final_tables_powerpoint_ready.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet, df in tables.items():
            if df.empty:
                warnings.append(f"Skipped table {sheet}: no source data.")
                continue
            df.to_excel(writer, sheet_name=sheet[:31], index=False)
            generated.append(sheet)

    if generated:
        format_workbook(workbook_path)
        add_manifest(manifest, workbook_path, "XLSX", "Tables", "Formatted PowerPoint-ready tables", "all")

    table_files = {
        "T1_Descriptive": ("table_1_descriptive_statistics", "Table 1 - Descriptive Statistics"),
        "T2_Stationarity": ("table_2_stationarity_tests", "Table 2 - Stationarity Tests"),
        "T3_Correlation": ("table_3_correlation_matrix", "Table 3 - Correlation Matrix"),
        "T4_VIF": ("table_4_vif", "Table 4 - Variance Inflation Factors"),
        "T5_QR_Full": ("table_5_qr_full_sample", "Table 5 - Full-Sample Quantile Regression Results"),
        "T6_Main_Findings": ("table_6_main_findings_summary", "Table 6 - Main Findings Summary"),
        "T7_Quantile_Diff": ("table_7_quantile_difference_tests", "Table 7 - Quantile-Difference Tests"),
        "T8_Diagnostics": ("table_8_diagnostics_summary", "Table 8 - Diagnostics Summary"),
        "T9_QLP_Full": ("table_9_qlp_full_sample", "Table 9 - QLP Full Sample Summary"),
        "T10_QLP_Sub_Summary": ("table_10_qlp_subsample_summary", "Table 10 - QLP Subsample Significance Summary"),
    }
    for sheet, df in tables.items():
        if df.empty:
            continue
        base, title = table_files[sheet]
        single_path = SUBDIRS["tables_excel"] / f"{base}.xlsx"
        df.to_excel(single_path, index=False)
        add_manifest(manifest, single_path, "XLSX", sheet, title, "all")
        png_path = SUBDIRS["tables_png"] / f"{base}.png"
        render_table_png(df, png_path, title, manifest, sheet, "all")

        # Compact PowerPoint copies for wide tables.
        if sheet in {"T5_QR_Full", "T9_QLP_Full"}:
            pp_path = SUBDIRS["tables_powerpoint_ready"] / f"{base}_compact.png"
            compact = compact_wide_table(df)
            render_table_png(compact, pp_path, f"{title} (Compact)", manifest, sheet, "slides")

    return generated


def compact_wide_table(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in df.columns if c.endswith("coef")]
    keep = [c for c in df.columns if c in {"Variable", "Horizon"}] + cols
    return df[keep].copy()


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor=COLORS["navy"].replace("#", ""))
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    diag_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    pos_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    neg_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = left if cell.column == 1 else center

        if ws.title == "T3_Correlation":
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row_name = ws.cell(row=r, column=1).value
                for c in range(2, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if headers[c - 1] == row_name:
                        cell.fill = diag_fill
                    else:
                        try:
                            val = float(cell.value)
                        except (TypeError, ValueError):
                            continue
                        if val > 0.5:
                            cell.fill = pos_fill
                        elif val < -0.5:
                            cell.fill = neg_fill

        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            max_len = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 38)

    wb.save(path)


def get_index_data(data: pd.DataFrame, warnings: List[str]) -> pd.DataFrame:
    required = ["Date", "STOXX 600 ESG-X", "STOXX 600"]
    if data.empty or not set(required).issubset(data.columns):
        warnings.append("Skipped index performance figures: missing Date/STOXX columns.")
        return pd.DataFrame()
    df = data[required].dropna().copy()
    df = df.sort_values("Date")
    for col in ["STOXX 600 ESG-X", "STOXX 600"]:
        df[col] = df[col] / df[col].iloc[0] * 100
    return df


def add_regime_shading(ax, y_top: Optional[float] = None, labels: bool = False) -> None:
    regimes = [
        ("2012-01-01", "2019-12-31", COLORS["very_light_gray"], 0.35, "Pre-crisis"),
        ("2020-01-01", "2022-12-31", COLORS["orange"], 0.08, "COVID & energy crisis"),
        ("2023-01-01", "2026-12-31", COLORS["green"], 0.08, "Post-crisis repricing"),
    ]
    for start, end, color, alpha, label in regimes:
        ax.axvspan(pd.Timestamp(start), pd.Timestamp(end), color=color, alpha=alpha, linewidth=0)
        if labels and y_top is not None:
            mid = pd.Timestamp(start) + (pd.Timestamp(end) - pd.Timestamp(start)) / 2
            ax.text(mid, y_top, label, ha="center", va="top", fontsize=8, color=COLORS["axis"])


def plot_index_figures(data: pd.DataFrame, manifest: List[dict], warnings: List[str]) -> None:
    df = get_index_data(data, warnings)
    if df.empty:
        return

    for regime in [False, True]:
        fig, ax = plt.subplots(figsize=(11, 6.2))
        if regime:
            add_regime_shading(ax, df[["STOXX 600 ESG-X", "STOXX 600"]].max().max() * 1.03, labels=True)
        ax.plot(df["Date"], df["STOXX 600 ESG-X"], color=COLORS["navy"], linewidth=2.0, label="STOXX 600 ESG-X")
        ax.plot(df["Date"], df["STOXX 600"], color=COLORS["gray"], linewidth=2.0, label="STOXX 600")
        format_axis(ax)
        ax.set_title("Normalized Performance of ESG-X and Conventional European Equity Benchmarks")
        ax.set_ylabel("Index level (Start = 100)")
        ax.set_xlabel("Date")
        ax.legend(frameon=False, loc="upper left", ncol=2)
        fig.tight_layout()
        base = "figure_1b_normalized_indices_regimes" if regime else "figure_1a_normalized_indices_clean"
        desc = "Normalized index performance with regime shading" if regime else "Normalized index performance"
        export_figure(fig, SUBDIRS["figures_main"], base, manifest, "Figure 1B" if regime else "Figure 1A", desc, "paper")
        plt.close(fig)


def construct_driver_data(data: pd.DataFrame, warnings: List[str]) -> pd.DataFrame:
    required = ["Date", "STOXX 600 ESG-X", "STOXX 600", "VSTOXX", "Bunds 10Y", "TTF", "Brent", "EUA"]
    if data.empty or not set(required).issubset(data.columns):
        warnings.append("Skipped driver time-series figures: required columns missing.")
        return pd.DataFrame()
    df = data[required].copy().dropna().sort_values("Date")
    df["spread"] = np.log(df["STOXX 600 ESG-X"] / df["STOXX 600 ESG-X"].shift(1)) - np.log(
        df["STOXX 600"] / df["STOXX 600"].shift(1)
    )
    return df


def plot_drivers(data: pd.DataFrame, manifest: List[dict], warnings: List[str]) -> None:
    df = construct_driver_data(data, warnings)
    if df.empty:
        return
    series = [
        ("VSTOXX", "VSTOXX"),
        ("Bunds 10Y", "Bund 10Y"),
        ("TTF", "TTF"),
        ("Brent", "Brent"),
        ("EUA", "EUA"),
        ("spread", "ESG Spread"),
    ]
    layouts = [((2, 3), "figure_2a_drivers_2x3", "Figure 2A"), ((3, 2), "figure_2b_drivers_3x2", "Figure 2B")]
    for (rows, cols), base, number in layouts:
        fig, axes = plt.subplots(rows, cols, figsize=(12, 7.2 if rows == 2 else 9.0), sharex=True)
        axes_flat = np.array(axes).flatten()
        for ax, (col, title) in zip(axes_flat, series):
            add_regime_shading(ax)
            ax.plot(df["Date"], df[col], color=COLORS["navy"], linewidth=1.6)
            format_axis(ax, zero_line=(col == "spread"))
            ax.set_title(title)
        for ax in axes_flat[len(series):]:
            ax.axis("off")
        fig.suptitle("Time-Series Evolution of the ESG Spread and Key Drivers", color=COLORS["navy"], fontsize=14)
        fig.tight_layout()
        export_figure(fig, SUBDIRS["figures_main"], base, manifest, number, "Drivers over time", "paper")
        plt.close(fig)


def draw_significant_marker(ax, x, y, color, is_sig, marker_size=42) -> None:
    if is_sig:
        ax.scatter(x, y, s=marker_size, facecolor=color, edgecolor=COLORS["gray"], linewidth=0.8, zorder=4)
        va = "bottom" if y >= 0 else "top"
        offset = 0.00003 if abs(y) < 0.001 else abs(y) * 0.08
        ax.text(x, y + (offset if y >= 0 else -offset), "*", ha="center", va=va, color=color, fontsize=10, fontweight="bold")
    else:
        ax.scatter(x, y, s=marker_size, facecolor="white", edgecolor=color, linewidth=1.3, zorder=4, alpha=0.9)


def plot_qr_panel(ax, df: pd.DataFrame, var: str, colored_markers: bool = False) -> None:
    sub = df[df["param"] == var].sort_values("quantile")
    if sub.empty:
        ax.axis("off")
        return
    x = sub["quantile"].to_numpy()
    y = sub["coef"].to_numpy()
    ax.fill_between(x, sub["ci_low"].to_numpy(), sub["ci_high"].to_numpy(), color=COLORS["light_gray"], alpha=0.20)
    if not colored_markers:
        ax.plot(x, y, color=COLORS["axis"], linewidth=2.0)
    else:
        ax.plot(x, y, color=COLORS["gray"], linewidth=1.6, alpha=0.75)
    for _, r in sub.iterrows():
        color = Q5_COLORS.get(round(float(r["quantile"]), 2), COLORS["axis"]) if colored_markers else COLORS["axis"]
        draw_significant_marker(ax, r["quantile"], r["coef"], color, significant(r))
    format_axis(ax, zero_line=True)
    ax.set_title(DISPLAY_NAMES[var])
    ax.set_xlabel("Quantile")
    ax.set_ylabel("Coefficient")
    ax.set_xticks([0.10, 0.25, 0.50, 0.75, 0.90])


def plot_qr_figures(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    df = sheets.get("quantile_reg", pd.DataFrame()).copy()
    if df.empty:
        warnings.append("Skipped QR figures: quantile_reg missing.")
        return

    for colored, base, number, out_dir in [
        (False, "figure_3a_qr_full_combined_single_line", "Figure 3A", SUBDIRS["figures_main"]),
        (True, "figure_3b_qr_full_combined_colored_quantiles", "Figure 3B", SUBDIRS["figures_variants"]),
    ]:
        fig, axes = plt.subplots(2, 3, figsize=(12.5, 7.2))
        axes_flat = axes.flatten()
        for ax, var in zip(axes_flat, MAIN_VARS):
            plot_qr_panel(ax, df, var, colored_markers=colored)
        axes_flat[-1].axis("off")
        axes_flat[-1].text(
            0.02,
            0.65,
            "Filled markers indicate 95% confidence intervals excluding zero.",
            color=COLORS["axis"],
            fontsize=10,
            wrap=True,
        )
        fig.suptitle("Quantile Regression Coefficients across the Conditional Distribution of the ESG Spread", color=COLORS["navy"], fontsize=14)
        fig.tight_layout()
        export_figure(fig, out_dir, base, manifest, number, "QR full-sample combined figure", "paper")
        plt.close(fig)

    for var in MAIN_VARS:
        fig, ax = plt.subplots(figsize=(10.5, 5.9))
        plot_qr_panel(ax, df, var, colored_markers=True)
        ax.set_title(f"Quantile Regression Coefficients: {DISPLAY_NAMES[var]}")
        fig.tight_layout()
        base = f"figure_3c_qr_{SLUGS[var]}"
        export_figure(fig, SUBDIRS["figures_variants"], base, manifest, "Figure 3C", f"Individual QR figure for {DISPLAY_NAMES[var]}", "paper")
        plt.close(fig)


def plot_rolling_panel(ax, df: pd.DataFrame, var: str) -> None:
    sub = df[df["param"] == var].copy()
    if sub.empty:
        ax.axis("off")
        return
    sub["end_date"] = pd.to_datetime(sub["end_date"], errors="coerce")
    add_regime_shading(ax)
    for q in [0.10, 0.50, 0.90]:
        sq = sub[np.isclose(sub["quantile"], q)].sort_values("end_date")
        ax.plot(sq["end_date"], sq["coef"], color=Q3_COLORS[q], linewidth=1.6, label=f"\u03c4={q:.2f}")
    format_axis(ax, zero_line=True)
    ax.set_title(DISPLAY_NAMES[var])
    ax.set_ylabel("Coefficient")


def plot_rolling_figures(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    for window, sheet, base, number in [
        (504, "rolling_504", "figure_4a_rolling_qr_504_all_variables", "Figure 4A"),
        (252, "rolling_252", "figure_4b_rolling_qr_252_all_variables", "Figure 4B"),
    ]:
        df = sheets.get(sheet, pd.DataFrame()).copy()
        if df.empty:
            warnings.append(f"Skipped rolling figure {window}: sheet {sheet} missing.")
            continue
        fig, axes = plt.subplots(2, 3, figsize=(12.5, 7.2), sharex=True)
        axes_flat = axes.flatten()
        for ax, var in zip(axes_flat, MAIN_VARS):
            plot_rolling_panel(ax, df, var)
        axes_flat[-1].axis("off")
        handles, labels = axes_flat[0].get_legend_handles_labels()
        axes_flat[-1].legend(handles, labels, frameon=False, loc="center")
        fig.suptitle(f"Rolling Quantile Regression Coefficients ({window}-Day Window)", color=COLORS["navy"], fontsize=14)
        fig.tight_layout()
        export_figure(fig, SUBDIRS["figures_main"], base, manifest, number, f"Rolling QR all variables {window}", "paper")
        plt.close(fig)

        for var in MAIN_VARS:
            fig, ax = plt.subplots(figsize=(10.5, 5.9))
            plot_rolling_panel(ax, df, var)
            ax.legend(frameon=False, ncol=3, loc="best")
            ax.set_xlabel("Window end date")
            ax.set_title(f"Rolling Quantile Coefficients ({window}-Day Window): {DISPLAY_NAMES[var]}")
            fig.tight_layout()
            name = f"figure_4c_rolling_{window}_{ROLLING_SLUGS[var]}"
            export_figure(fig, SUBDIRS["figures_appendix"], name, manifest, "Figure 4C", f"Individual rolling QR for {DISPLAY_NAMES[var]}", "appendix")
            plt.close(fig)


def plot_qlp_panel(ax, df: pd.DataFrame, var: str, errorbar: bool = False) -> None:
    sub = df[df["param"] == var].copy()
    if sub.empty:
        ax.axis("off")
        return
    for q in [0.10, 0.50, 0.90]:
        sq = sub[np.isclose(sub["quantile"], q)].sort_values("horizon")
        if sq.empty:
            continue
        x = sq["horizon"].to_numpy()
        y = sq["coef"].to_numpy()
        color = Q3_COLORS[q]
        if errorbar:
            yerr = np.vstack([y - sq["ci_low"].to_numpy(), sq["ci_high"].to_numpy() - y])
            ax.errorbar(x, y, yerr=yerr, color=color, linewidth=1.8, marker="o", capsize=3, label=f"\u03c4={q:.2f}", alpha=0.9)
        else:
            ax.plot(x, y, color=color, linewidth=2.0, marker="o", label=f"\u03c4={q:.2f}")
            ax.fill_between(x, sq["ci_low"].to_numpy(), sq["ci_high"].to_numpy(), color=color, alpha=0.15)
        for _, r in sq.iterrows():
            draw_significant_marker(ax, r["horizon"], r["coef"], color, significant(r), marker_size=38)
    format_axis(ax, zero_line=True)
    ax.set_title(DISPLAY_NAMES[var])
    ax.set_xlabel("Horizon")
    ax.set_ylabel("QLP coefficient")
    ax.set_xticks(HORIZONS)


def plot_qlp_full(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    df = sheets.get("qlp_key_results", pd.DataFrame()).copy()
    if df.empty:
        warnings.append("Skipped QLP full figures: qlp_key_results missing.")
        return
    full = df[df["sample_label"] == "full"].copy()
    if full.empty:
        warnings.append("Skipped QLP full figures: no full-sample rows.")
        return

    for errorbar, base, number in [
        (False, "figure_5a_qlp_full_all_variables_ribbon", "Figure 5A"),
        (True, "figure_5b_qlp_full_all_variables_errorbar", "Figure 5B"),
    ]:
        fig, axes = plt.subplots(2, 3, figsize=(12.5, 7.2))
        axes_flat = axes.flatten()
        for ax, var in zip(axes_flat, MAIN_VARS):
            plot_qlp_panel(ax, full, var, errorbar=errorbar)
        axes_flat[-1].axis("off")
        handles, labels = axes_flat[0].get_legend_handles_labels()
        axes_flat[-1].legend(handles, labels, frameon=False, loc="center")
        axes_flat[-1].text(0.02, 0.25, "Filled markers indicate 95% confidence intervals excluding zero.", fontsize=10, color=COLORS["axis"], wrap=True)
        fig.suptitle("Dynamic Quantile Local Projection Responses of the ESG Spread", color=COLORS["navy"], fontsize=14)
        fig.tight_layout()
        out_dir = SUBDIRS["figures_main"] if not errorbar else SUBDIRS["figures_variants"]
        export_figure(fig, out_dir, base, manifest, number, "QLP full-sample all variables", "paper")
        plt.close(fig)

    for var in MAIN_VARS:
        fig, ax = plt.subplots(figsize=(10.5, 5.9))
        plot_qlp_panel(ax, full, var, errorbar=False)
        ax.legend(frameon=False, ncol=3, loc="best")
        ax.set_title(f"Dynamic QLP Responses: {DISPLAY_NAMES[var]}")
        fig.tight_layout()
        base = f"figure_5c_qlp_full_{SLUGS[var]}"
        export_figure(fig, SUBDIRS["figures_variants"], base, manifest, "Figure 5C", f"Individual QLP full-sample figure for {DISPLAY_NAMES[var]}", "paper")
        plt.close(fig)


def plot_qlp_subsamples(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    df = sheets.get("qlp_key_results", pd.DataFrame()).copy()
    if df.empty:
        warnings.append("Skipped QLP subsample figures: qlp_key_results missing.")
        return
    sub = df[df["sample_label"].isin(SUBSAMPLES)].copy()
    if sub.empty:
        warnings.append("Skipped QLP subsample figures: no subsample rows.")
        return

    for var in MAIN_VARS:
        fig, axes = plt.subplots(1, 3, figsize=(14, 4.7), sharey=True)
        for ax, sample in zip(axes, SUBSAMPLES):
            ss = sub[(sub["sample_label"] == sample) & (sub["param"] == var)]
            plot_qlp_panel(ax, ss, var, errorbar=False)
            ax.set_title(sample_label(sample))
        axes[0].set_ylabel("QLP coefficient")
        handles, labels = axes[0].get_legend_handles_labels()
        if handles:
            axes[-1].legend(handles, labels, frameon=False, loc="best")
        fig.suptitle(f"Subsample Evidence: QLP Responses for {DISPLAY_NAMES[var]}", color=COLORS["navy"], fontsize=14)
        fig.tight_layout()
        base = f"figure_6a_qlp_subsamples_{SLUGS[var]}"
        export_figure(fig, SUBDIRS["figures_appendix"], base, manifest, "Figure 6A", f"QLP subsample evidence for {DISPLAY_NAMES[var]}", "appendix")
        plt.close(fig)

        # Compact slide version: median quantile only.
        fig, ax = plt.subplots(figsize=(10.5, 5.9))
        for sample in SUBSAMPLES:
            ss = sub[(sub["sample_label"] == sample) & (sub["param"] == var) & np.isclose(sub["quantile"], 0.50)].sort_values("horizon")
            ax.plot(ss["horizon"], ss["coef"], marker="o", linewidth=2.0, label=sample_label(sample))
        format_axis(ax, zero_line=True)
        ax.set_xticks(HORIZONS)
        ax.set_xlabel("Horizon")
        ax.set_ylabel("QLP coefficient")
        ax.set_title(f"Subsample Median QLP Responses: {DISPLAY_NAMES[var]}")
        ax.legend(frameon=False, ncol=3, loc="best")
        fig.tight_layout()
        base = f"figure_6b_qlp_subsamples_compact_{SLUGS[var]}"
        export_figure(fig, SUBDIRS["figures_slides"], base, manifest, "Figure 6B", f"Compact slide QLP subsample figure for {DISPLAY_NAMES[var]}", "slides")
        plt.close(fig)


def heatmap_colors(values: List[int]) -> List[str]:
    return [COLORS["green"] if v > 0 else COLORS["burgundy"] if v < 0 else COLORS["very_light_gray"] for v in values]


def plot_qr_subsample_heatmap(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    qr_subsample_quantiles = [0.10, 0.25, 0.50, 0.75, 0.90]
    rows = []
    columns = []
    for sample in SUBSAMPLES:
        sheet = f"qr_{sample}"
        df = sheets.get(sheet, pd.DataFrame()).copy()
        if df.empty:
            continue
        for q in qr_subsample_quantiles:
            columns.append(f"{sample_label(sample)}\n{q_label(q)}")
        for var in MAIN_VARS:
            pass
    columns = list(dict.fromkeys(columns))
    if not columns:
        warnings.append("Skipped Figure 7: QR subsample sheets missing.")
        return
    matrix = []
    for var in MAIN_VARS:
        row_vals = []
        for sample in SUBSAMPLES:
            df = sheets.get(f"qr_{sample}", pd.DataFrame()).copy()
            for q in qr_subsample_quantiles:
                if df.empty:
                    row_vals.append(0)
                    continue
                ss = df[(df["param"] == var) & np.isclose(df["quantile"], q)]
                if ss.empty or not significant(ss.iloc[0]):
                    row_vals.append(0)
                else:
                    row_vals.append(1 if ss["coef"].iloc[0] > 0 else -1)
        matrix.append(row_vals)

    draw_heatmap(matrix, [DISPLAY_NAMES[v] for v in MAIN_VARS], columns, "QR Subsample Significance Summary", SUBDIRS["figures_appendix"], "figure_7_qr_subsample_significance_heatmap", manifest, "Figure 7", "QR subsample significance heatmap across all estimated quantiles")


def plot_qlp_significance_heatmap(sheets: Dict[str, pd.DataFrame], manifest: List[dict], warnings: List[str]) -> None:
    df = sheets.get("qlp_key_results", pd.DataFrame()).copy()
    if df.empty:
        warnings.append("Skipped Figure 8: qlp_key_results missing.")
        return
    columns = [f"h{h}\n{q_label(q)}" for h in HORIZONS for q in [0.10, 0.50, 0.90]]
    fig, axes = plt.subplots(2, 2, figsize=(14, 8.5))
    for ax, sample in zip(axes.flatten(), ["full"] + SUBSAMPLES):
        matrix = []
        for var in MAIN_VARS:
            row_vals = []
            for h in HORIZONS:
                for q in [0.10, 0.50, 0.90]:
                    ss = df[(df["sample_label"] == sample) & (df["param"] == var) & (df["horizon"] == h) & np.isclose(df["quantile"], q)]
                    if ss.empty or not significant(ss.iloc[0]):
                        row_vals.append(0)
                    else:
                        row_vals.append(1 if ss["coef"].iloc[0] > 0 else -1)
            matrix.append(row_vals)
        draw_heatmap_on_axis(ax, matrix, [DISPLAY_NAMES[v] for v in MAIN_VARS], columns, sample_label(sample))
    fig.suptitle("QLP Significance Heatmap", color=COLORS["navy"], fontsize=14)
    fig.tight_layout()
    export_figure(fig, SUBDIRS["figures_appendix"], "figure_8_qlp_significance_heatmap", manifest, "Figure 8", "QLP significance heatmap", "appendix")
    plt.close(fig)


def draw_heatmap(
    matrix: List[List[int]],
    ylabels: List[str],
    xlabels: List[str],
    title: str,
    out_dir: Path,
    basename: str,
    manifest: List[dict],
    number: str,
    description: str,
) -> None:
    fig_width = max(13, 3.5 + 0.85 * len(xlabels))
    fig, ax = plt.subplots(figsize=(fig_width, 5.5))
    draw_heatmap_on_axis(ax, matrix, ylabels, xlabels, title)
    fig.tight_layout()
    export_figure(fig, out_dir, basename, manifest, number, description, "appendix")
    plt.close(fig)


def draw_heatmap_on_axis(ax, matrix: List[List[int]], ylabels: List[str], xlabels: List[str], title: str) -> None:
    color_map = {1: COLORS["green"], -1: COLORS["burgundy"], 0: COLORS["very_light_gray"]}
    arr = np.array(matrix)
    for i in range(arr.shape[0]):
        for j in range(arr.shape[1]):
            val = int(arr[i, j])
            rect = patches.Rectangle((j, i), 1, 1, facecolor=color_map[val], edgecolor="white", linewidth=1.2)
            ax.add_patch(rect)
            txt = "+" if val > 0 else "-" if val < 0 else ""
            ax.text(j + 0.5, i + 0.5, txt, ha="center", va="center", color="white" if val != 0 else COLORS["gray"], fontweight="bold")
    ax.set_xlim(0, arr.shape[1])
    ax.set_ylim(0, arr.shape[0])
    ax.set_xticks(np.arange(arr.shape[1]) + 0.5)
    ax.set_yticks(np.arange(arr.shape[0]) + 0.5)
    ax.set_xticklabels(xlabels, rotation=45, ha="right", fontsize=8)
    ax.set_yticklabels(ylabels, fontsize=9)
    ax.invert_yaxis()
    ax.set_title(title, color=COLORS["navy"])
    ax.tick_params(length=0)
    for spine in ax.spines.values():
        spine.set_visible(False)


def draw_workflow_box(ax, xy, width, height, title, lines, accent=COLORS["navy"]):
    x, y = xy
    box = patches.FancyBboxPatch(
        (x, y),
        width,
        height,
        boxstyle="round,pad=0.015,rounding_size=0.02",
        linewidth=1.2,
        edgecolor=accent,
        facecolor="white",
    )
    ax.add_patch(box)
    ax.add_patch(patches.Rectangle((x, y + height - 0.16), width, 0.16, color=accent, linewidth=0))
    ax.text(x + width / 2, y + height - 0.08, title, ha="center", va="center", color="white", fontweight="bold", fontsize=9)
    body = "\n".join(lines)
    ax.text(x + 0.03, y + height - 0.22, body, ha="left", va="top", fontsize=8.1, color=COLORS["axis"])


def arrow(ax, start, end):
    ax.annotate(
        "",
        xy=end,
        xytext=start,
        arrowprops=dict(arrowstyle="->", color=COLORS["gray"], linewidth=1.5, shrinkA=4, shrinkB=4),
    )


def plot_methodology(manifest: List[dict]) -> None:
    fig, ax = plt.subplots(figsize=(14, 7.875))
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    draw_workflow_box(ax, (0.03, 0.62), 0.18, 0.27, "Data", ["STOXX 600 ESG-X", "STOXX 600", "VSTOXX", "Bund 10Y", "TTF, Brent, EUA"], COLORS["navy"])
    draw_workflow_box(ax, (0.27, 0.62), 0.20, 0.27, "Variable Construction", ["ESG Spread = r(ESG-X) - r(STOXX 600)", "First differences: VSTOXX, Bund", "Log returns: TTF, Brent, EUA"], COLORS["navy"])
    draw_workflow_box(ax, (0.53, 0.62), 0.18, 0.27, "Preliminary Checks", ["ADF / PP tests", "Descriptive statistics", "Correlation matrix", "VIF"], COLORS["green"])
    draw_workflow_box(ax, (0.78, 0.62), 0.19, 0.27, "Baseline Model", ["Quantile Regression", "\u03c4 = 0.10, 0.25, 0.50, 0.75, 0.90", "Moving block bootstrap"], COLORS["burgundy"])
    draw_workflow_box(ax, (0.18, 0.20), 0.27, 0.27, "Regime and Stability Analysis", ["Subsamples: 2012-2019, 2020-2022, 2023-2026", "Rolling QR: 252 and 504 days"], COLORS["green"])
    draw_workflow_box(ax, (0.55, 0.20), 0.22, 0.27, "Dynamic Extension", ["Quantile Local Projections", "\u03c4 = 0.10, 0.50, 0.90", "h = 1, 5, 10, 20 days"], COLORS["navy"])
    draw_workflow_box(ax, (0.80, 0.20), 0.18, 0.27, "Economic Interpretation", ["Energy-exclusion wedge", "Market-stress channel", "European gas shock channel", "ESG benchmark risk implications"], COLORS["burgundy"])
    arrow(ax, (0.21, 0.755), (0.27, 0.755))
    arrow(ax, (0.47, 0.755), (0.53, 0.755))
    arrow(ax, (0.71, 0.755), (0.78, 0.755))
    arrow(ax, (0.875, 0.62), (0.32, 0.47))
    arrow(ax, (0.875, 0.62), (0.66, 0.47))
    arrow(ax, (0.77, 0.335), (0.80, 0.335))
    fig.suptitle("Empirical Strategy and Methodological Workflow", color=COLORS["navy"], fontsize=16, fontweight="bold")
    fig.tight_layout()
    export_figure(fig, SUBDIRS["methodology_flow"], "figure_9_methodological_workflow", manifest, "Figure 9", "Methodological workflow", "paper")
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(13.33, 7.5))
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    boxes = [
        ((0.06, 0.38), "Data", ["Market, volatility, rates", "energy and carbon series"], COLORS["navy"]),
        ((0.30, 0.38), "ESG Spread & Drivers", ["Spread construction", "returns and differences"], COLORS["green"]),
        ((0.54, 0.38), "QR + Bootstrap", ["Distributional estimates", "block bootstrap inference"], COLORS["burgundy"]),
        ((0.78, 0.38), "Robustness", ["Subsamples", "Rolling QR", "QLP dynamics"], COLORS["navy"]),
    ]
    for xy, title, lines, color in boxes:
        draw_workflow_box(ax, xy, 0.18, 0.25, title, lines, color)
    arrow(ax, (0.24, 0.505), (0.30, 0.505))
    arrow(ax, (0.48, 0.505), (0.54, 0.505))
    arrow(ax, (0.72, 0.505), (0.78, 0.505))
    fig.suptitle("Empirical Strategy", color=COLORS["navy"], fontsize=18, fontweight="bold")
    fig.tight_layout()
    export_figure(fig, SUBDIRS["figures_slides"], "figure_10_empirical_strategy_slide", manifest, "Figure 10", "One-slide empirical strategy", "slides")
    plt.close(fig)


def write_summary(
    manifest: List[dict],
    generated_tables: List[str],
    used_sheets: List[str],
    warnings: List[str],
) -> None:
    manifest_df = pd.DataFrame(manifest)
    manifest_path = OUTPUT_DIR / "output_manifest.csv"
    manifest_df.to_csv(manifest_path, index=False)

    summary_path = OUTPUT_DIR / "visual_outputs_summary.txt"
    figures = manifest_df[manifest_df["output_type"].isin(["PNG", "PDF", "SVG"])] if not manifest_df.empty else pd.DataFrame()
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("Final Visual Package Summary\n")
        f.write("============================\n\n")
        f.write(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Input data: {DATA_FILE}\n")
        f.write(f"Input results: {RESULTS_FILE}\n")
        f.write(f"Output folder: {OUTPUT_DIR}\n\n")
        f.write("Input sheets used:\n")
        for s in used_sheets:
            f.write(f"- {s}\n")
        f.write("\nTables generated:\n")
        for t in generated_tables:
            f.write(f"- {t}\n")
        f.write("\nFigures generated:\n")
        for _, row in figures.iterrows():
            f.write(f"- {row['output_file']}\n")
        f.write("\nNotes:\n")
        f.write("- Filled markers indicate 95% confidence intervals excluding zero.\n")
        f.write("- Hollow markers indicate coefficients whose 95% confidence intervals include zero.\n")
        f.write("\nWarnings / skipped outputs:\n")
        if warnings:
            for w in warnings:
                f.write(f"- {w}\n")
        else:
            f.write("- None\n")


def main() -> None:
    configure_style()
    ensure_dirs()

    warnings: List[str] = []
    manifest: List[dict] = []

    sheets, used_sheets = read_sheets(warnings)
    data = read_data(warnings)

    tables = build_tables(sheets)
    generated_tables = export_tables(tables, manifest, warnings)

    plot_index_figures(data, manifest, warnings)
    plot_drivers(data, manifest, warnings)
    plot_qr_figures(sheets, manifest, warnings)
    plot_rolling_figures(sheets, manifest, warnings)
    plot_qlp_full(sheets, manifest, warnings)
    plot_qlp_subsamples(sheets, manifest, warnings)
    plot_qr_subsample_heatmap(sheets, manifest, warnings)
    plot_qlp_significance_heatmap(sheets, manifest, warnings)
    plot_methodology(manifest)

    write_summary(manifest, generated_tables, used_sheets, warnings)

    fig_count = sum(1 for item in manifest if item["output_type"] in {"PNG", "PDF", "SVG"} and item["figure_or_table_number"].startswith("Figure"))
    table_count = len(generated_tables)
    print("Final visual package completed.")
    print(f"Figures generated: {fig_count}")
    print(f"Tables generated: {table_count}")
    print(f"Output folder: {OUTPUT_DIR}")
    if warnings:
        print("Skipped outputs / warnings:")
        for w in warnings:
            print(f"- {w}")
    else:
        print("Skipped outputs / warnings: None")


if __name__ == "__main__":
    main()
